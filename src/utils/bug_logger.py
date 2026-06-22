"""Deterministic bug logger -> Excel workbook for the SmileCareMedicine QA framework.

Used by the `bug-house` pipeline agent (stage 6) to persist confirmed defects into
`bug_reports/bug_house.xlsx` in a standard, senior-QA bug-report format: one worksheet
per module plus a rolled-up ``Summary`` sheet.

Bugs are keyed by a stable **Bug ID** and *upserted* — re-running a module updates the
existing row (``Last Seen`` / ``Occurrences`` / ``Status``) instead of creating
duplicates, so the workbook is a cumulative bug *memory* that grows across every
execution rather than being overwritten.

Why a helper instead of letting the agent write the .xlsx directly: writing binary
Excel by hand corrupts easily. Routing every write through this module keeps the
format consistent, deduped, and auditable.

CLI (what the bug-house agent calls):
    python -m src.utils.bug_logger --add-many bugs.json     # list[dict] of bugs
    python -m src.utils.bug_logger --add bug.json           # single bug dict
    python -m src.utils.bug_logger --summary                # rebuild Summary sheet only
"""

from __future__ import annotations

import argparse
import json
from datetime import datetime, timezone
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

# --------------------------------------------------------------------------- #
# Format definition (the standard senior-QA bug report)
# --------------------------------------------------------------------------- #
WORKBOOK_PATH = Path(__file__).resolve().parents[2] / "bug_reports" / "bug_house.xlsx"

COLUMNS: list[str] = [
    "Bug ID",            # stable key, e.g. BUG-AUTH-001
    "Module",            # auth | products | cart | ...
    "Title",             # one-line summary
    "Severity",          # Critical | Major | Minor | Trivial
    "Priority",          # P1 | P2 | P3 | P4
    "Status",            # New | Open | Reopened | Retest | Closed | Deferred
    "Test Case ID",      # traceability link, e.g. AUTH-007
    "Environment",       # browser / OS / URL
    "Steps to Reproduce",
    "Test Data",
    "Expected Result",
    "Actual Result",
    "Evidence",          # path(s) to trace/screenshot/video in reports/
    "First Seen",
    "Last Seen",
    "Occurrences",
    "Reported By",
    "Notes",
]

SEVERITY_ORDER = ["Critical", "Major", "Minor", "Trivial"]

# styling
_HEADER_FILL = PatternFill("solid", fgColor="1F3864")
_HEADER_FONT = Font(bold=True, color="FFFFFF")
_SEVERITY_FILL = {
    "Critical": PatternFill("solid", fgColor="C00000"),
    "Major": PatternFill("solid", fgColor="ED7D31"),
    "Minor": PatternFill("solid", fgColor="FFD966"),
    "Trivial": PatternFill("solid", fgColor="A9D18E"),
}
_SUMMARY_SHEET = "Summary"


def _now() -> str:
    """UTC timestamp; uses real wall-clock (this runs in normal Python, not a workflow)."""
    return datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M:%S UTC")


def _module_sheet_name(module: str) -> str:
    # Excel sheet names: <=31 chars, no : \ / ? * [ ]
    safe = "".join(c for c in module if c not in r":\/?*[]").strip() or "misc"
    return safe[:31]


def _ensure_workbook() -> Workbook:
    WORKBOOK_PATH.parent.mkdir(parents=True, exist_ok=True)
    if WORKBOOK_PATH.exists():
        return load_workbook(WORKBOOK_PATH)
    wb = Workbook()
    # rename the default sheet into the Summary sheet
    wb.active.title = _SUMMARY_SHEET
    return wb


def _ensure_module_sheet(wb: Workbook, module: str) -> Worksheet:
    name = _module_sheet_name(module)
    if name in wb.sheetnames:
        return wb[name]
    ws = wb.create_sheet(title=name)
    ws.append(COLUMNS)
    for col_idx in range(1, len(COLUMNS) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = Alignment(vertical="center", wrap_text=True)
    ws.freeze_panes = "A2"
    _autosize(ws)
    return ws


def _autosize(ws: Worksheet) -> None:
    widths = {
        "Bug ID": 16, "Module": 12, "Title": 40, "Severity": 11, "Priority": 9,
        "Status": 11, "Test Case ID": 14, "Environment": 24, "Steps to Reproduce": 50,
        "Test Data": 24, "Expected Result": 40, "Actual Result": 40, "Evidence": 40,
        "First Seen": 22, "Last Seen": 22, "Occurrences": 12, "Reported By": 16, "Notes": 30,
    }
    for idx, col in enumerate(COLUMNS, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = widths.get(col, 18)


def _find_row(ws: Worksheet, bug_id: str) -> int | None:
    for row in range(2, ws.max_row + 1):
        if str(ws.cell(row=row, column=1).value or "").strip() == bug_id:
            return row
    return None


def _normalize(bug: dict) -> dict:
    """Fill defaults and coerce keys to the canonical column names."""
    b = {k: ("" if v is None else v) for k, v in bug.items()}
    b.setdefault("Severity", "Major")
    b.setdefault("Priority", "P2")
    b.setdefault("Status", "New")
    b.setdefault("Reported By", "bug-house-agent")
    if not b.get("Bug ID"):
        raise ValueError("bug is missing required 'Bug ID'")
    if not b.get("Module"):
        raise ValueError(f"bug {b['Bug ID']} is missing required 'Module'")
    if b["Severity"] not in SEVERITY_ORDER:
        b["Severity"] = "Major"
    return b


def upsert_bug(bug: dict, wb: Workbook | None = None) -> str:
    """Insert a new bug or update an existing one (matched on Bug ID).

    Returns one of: "added", "updated", "reopened".
    """
    own = wb is None
    wb = wb or _ensure_workbook()
    b = _normalize(bug)
    ws = _ensure_module_sheet(wb, b["Module"])
    now = _now()
    existing = _find_row(ws, b["Bug ID"])

    if existing is None:
        action = "added"
        b["First Seen"] = now
        b["Last Seen"] = now
        b["Occurrences"] = 1
        ws.append([b.get(col, "") for col in COLUMNS])
        _style_row(ws, ws.max_row, b["Severity"])
    else:
        action = "updated"
        prev_status = str(ws.cell(row=existing, column=COLUMNS.index("Status") + 1).value or "")
        prev_occ = ws.cell(row=existing, column=COLUMNS.index("Occurrences") + 1).value or 0
        # Re-appearing after Closed => reopen
        new_status = b["Status"]
        if prev_status == "Closed" and b["Status"] not in ("Closed", "Deferred"):
            new_status = "Reopened"
            action = "reopened"
        # update the volatile fields; keep First Seen
        updates = {
            "Title": b.get("Title", ""),
            "Severity": b["Severity"],
            "Priority": b["Priority"],
            "Status": new_status,
            "Test Case ID": b.get("Test Case ID", ""),
            "Environment": b.get("Environment", ""),
            "Steps to Reproduce": b.get("Steps to Reproduce", ""),
            "Test Data": b.get("Test Data", ""),
            "Expected Result": b.get("Expected Result", ""),
            "Actual Result": b.get("Actual Result", ""),
            "Evidence": b.get("Evidence", ""),
            "Last Seen": now,
            "Occurrences": int(prev_occ) + 1,
            "Notes": b.get("Notes", ""),
        }
        for col, val in updates.items():
            if val == "" and col not in ("Last Seen", "Occurrences", "Status", "Severity", "Priority"):
                continue  # don't blank out previously-captured detail with empty re-reports
            ws.cell(row=existing, column=COLUMNS.index(col) + 1, value=val)
        _style_row(ws, existing, new_status if new_status in _SEVERITY_FILL else b["Severity"])

    if own:
        refresh_summary(wb)
        wb.save(WORKBOOK_PATH)
    return action


def _style_row(ws: Worksheet, row: int, severity: str) -> None:
    sev_cell = ws.cell(row=row, column=COLUMNS.index("Severity") + 1)
    fill = _SEVERITY_FILL.get(severity)
    if fill:
        sev_cell.fill = fill
        sev_cell.font = Font(bold=True)
    for col in range(1, len(COLUMNS) + 1):
        ws.cell(row=row, column=col).alignment = Alignment(vertical="top", wrap_text=True)


def refresh_summary(wb: Workbook) -> None:
    """Rebuild the Summary sheet: per-module counts and severity breakdown."""
    if _SUMMARY_SHEET in wb.sheetnames:
        del wb[_SUMMARY_SHEET]
    ws = wb.create_sheet(title=_SUMMARY_SHEET, index=0)

    ws["A1"] = "SmileCareMedicine — Bug House"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A2"] = f"Last updated: {_now()}"
    ws["A2"].font = Font(italic=True, color="808080")

    header = ["Module", "Total", *SEVERITY_ORDER, "Open", "Closed"]
    ws.append([])  # row 3 spacer
    ws.append(header)
    head_row = ws.max_row
    for col_idx in range(1, len(header) + 1):
        c = ws.cell(row=head_row, column=col_idx)
        c.fill = _HEADER_FILL
        c.font = _HEADER_FONT

    grand = {"Total": 0, "Open": 0, "Closed": 0, **{s: 0 for s in SEVERITY_ORDER}}
    for name in wb.sheetnames:
        if name == _SUMMARY_SHEET:
            continue
        sheet = wb[name]
        counts = {"Total": 0, "Open": 0, "Closed": 0, **{s: 0 for s in SEVERITY_ORDER}}
        sev_col = COLUMNS.index("Severity") + 1
        status_col = COLUMNS.index("Status") + 1
        for row in range(2, sheet.max_row + 1):
            if not sheet.cell(row=row, column=1).value:
                continue
            counts["Total"] += 1
            sev = str(sheet.cell(row=row, column=sev_col).value or "")
            if sev in counts:
                counts[sev] += 1
            status = str(sheet.cell(row=row, column=status_col).value or "")
            if status == "Closed":
                counts["Closed"] += 1
            else:
                counts["Open"] += 1
        if counts["Total"] == 0:
            continue
        ws.append([name, counts["Total"], *[counts[s] for s in SEVERITY_ORDER],
                   counts["Open"], counts["Closed"]])
        for k in grand:
            grand[k] += counts[k]

    total_row = ws.append(["TOTAL", grand["Total"], *[grand[s] for s in SEVERITY_ORDER],
                           grand["Open"], grand["Closed"]])
    for col_idx in range(1, len(header) + 1):
        ws.cell(row=ws.max_row, column=col_idx).font = Font(bold=True)

    for idx in range(1, len(header) + 1):
        ws.column_dimensions[get_column_letter(idx)].width = 14
    ws.column_dimensions["A"].width = 18


def add_many(bugs: list[dict]) -> dict:
    """Upsert a batch of bugs in one workbook open/save. Returns action tally."""
    wb = _ensure_workbook()
    tally = {"added": 0, "updated": 0, "reopened": 0}
    for bug in bugs:
        tally[upsert_bug(bug, wb=wb)] += 1
    refresh_summary(wb)
    wb.save(WORKBOOK_PATH)
    return tally


# --------------------------------------------------------------------------- #
# CLI
# --------------------------------------------------------------------------- #
def _main() -> None:
    parser = argparse.ArgumentParser(description="Append/upsert bugs into bug_house.xlsx")
    g = parser.add_mutually_exclusive_group(required=True)
    g.add_argument("--add", metavar="JSON", help="path to a JSON file with ONE bug dict")
    g.add_argument("--add-many", metavar="JSON", help="path to a JSON file with a LIST of bug dicts")
    g.add_argument("--summary", action="store_true", help="rebuild the Summary sheet only")
    args = parser.parse_args()

    if args.summary:
        wb = _ensure_workbook()
        refresh_summary(wb)
        wb.save(WORKBOOK_PATH)
        print(f"Summary refreshed -> {WORKBOOK_PATH}")
        return

    path = Path(args.add or args.add_many)
    data = json.loads(path.read_text(encoding="utf-8"))
    bugs = data if isinstance(data, list) else [data]
    tally = add_many(bugs)
    print(f"Bug House updated -> {WORKBOOK_PATH}")
    print(f"  added={tally['added']}  updated={tally['updated']}  reopened={tally['reopened']}")


if __name__ == "__main__":
    _main()
